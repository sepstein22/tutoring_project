"""Loading FACED files, and converting subjects to a common unit."""
from __future__ import annotations

import glob
import os
import pickle
import unicodedata
from dataclasses import dataclass

import numpy as np

from . import config as C


# ---------------------------------------------------------------- discovery

def _first_existing(root, candidates, pattern="sub*"):
    """First candidate directory that exists and holds matching files.

    An empty directory does not count, so a partial download cannot be
    mistaken for a complete one.
    """
    for name in candidates:
        path = os.path.join(root, *name.split("/"))
        if os.path.isdir(path) and glob.glob(os.path.join(path, pattern)):
            return path
    return None


def locate(root):
    """Find the pieces of the release under `root`.

    Tolerates the layout variants listed in config.DIRS.
    """
    found = {key: _first_existing(root, names) for key, names in C.DIRS.items()}
    for key, name in C.FILES.items():
        hits = glob.glob(os.path.join(root, "**", name), recursive=True)
        found[key] = hits[0] if hits else None
    return found


def canonical_subject_id(path):
    """Subject index parsed from a filename: sub007_rating.mat -> 7.

    Uses the first run of digits, so a suffix such as "_v2" cannot change the
    identity.

    Raises
        ValueError if the filename holds no digits.
    """
    stem = os.path.basename(path)
    digits = ""
    for ch in stem:
        if ch.isdigit():
            digits += ch
        elif digits:
            break
    if not digits:
        raise ValueError("no subject number in filename: %s" % stem)
    return int(digits)


def index_by_subject(paths, what="file"):
    """Map subject id to path.

    Raises
        ValueError if two paths resolve to the same subject.
    """
    out = {}
    for path in paths:
        key = canonical_subject_id(path)
        if key in out:
            raise ValueError("two %s candidates for subject %03d:\n  %s\n  %s"
                             % (what, key, out[key], path))
        out[key] = path
    return out


def subject_files(directory, pattern="sub*.pkl*"):
    """Subject files ordered by the integer in the filename, not lexically."""
    paths = [p for p in glob.glob(os.path.join(directory, pattern))
             if os.path.isfile(p)]
    if not paths:
        raise FileNotFoundError("no %r under %s" % (pattern, directory))

    return sorted(paths, key=canonical_subject_id)


# ------------------------------------------------------------ recording info

@dataclass(frozen=True)
class Recording:
    """One row of Recording_info.csv."""

    subject: str
    gender: str
    age: float
    cohort: int
    sample_rate: int
    unit: str

    @property
    def index(self):
        """Integer parsed from the subject label."""
        return int("".join(c for c in self.subject if c.isdigit()))


def load_recording_info(path):
    """Parse Recording_info.csv into {subject_index: Recording}.

    Headers are matched case- and whitespace-insensitively; the released file
    has a trailing space in "Cohort ".
    """
    import csv

    with open(path, newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        raise ValueError("%s is empty" % path)

    def pick(row, *wanted):
        for key, value in row.items():
            if key is None:
                continue
            if key.strip().lower() in wanted:
                return value.strip() if isinstance(value, str) else value
        raise KeyError("no column matching %s in %s" % (wanted, list(row)))

    out = {}
    for row in rows:
        rec = Recording(
            subject=pick(row, "sub", "subject"),
            gender=pick(row, "gender", "sex"),
            age=float(pick(row, "age")),
            cohort=int(pick(row, "cohort")),
            sample_rate=int(float(pick(row, "sample_rate", "samplerate",
                                       "fs"))),
            unit=pick(row, "unit", "units"),
        )
        out[rec.index] = rec

    if len(out) != len(rows):
        raise ValueError("duplicate subject indices in %s" % path)
    return out


def unit_summary(info):
    """Counts of unit, cohort and sampling rate, and their cross-tabulation."""
    from collections import Counter
    combos = Counter((r.cohort, r.sample_rate, r.unit) for r in info.values())
    return {
        "n_subject": len(info),
        "unit": dict(Counter(r.unit for r in info.values())),
        "cohort": dict(Counter(r.cohort for r in info.values())),
        "sample_rate": dict(Counter(r.sample_rate for r in info.values())),
        "combinations": {"cohort=%s sfreq=%s unit=%s" % k: v
                         for k, v in sorted(combos.items())},
        "non_canonical": sorted(i for i, r in info.items()
                                if r.unit != C.CANONICAL_UNIT),
    }


# --------------------------------------------------------------------- units

def normalize_unit(unit):
    """Fold any spelling of a unit onto its canonical form.

    Handles both characters that render as mu: U+00B5 MICRO SIGN and U+03BC
    GREEK SMALL LETTER MU, which are not equal.

    Raises
        KeyError on an unrecognised unit.
    """
    value = unicodedata.normalize("NFKC", str(unit)).strip()
    value = value.replace("µ", "u").replace("μ", "u").lower()
    if value not in C.UNIT_ALIASES:
        raise KeyError("unknown unit %r; accepted forms include V, uV, µV, "
                       "μV, microvolts" % (unit,))
    return C.UNIT_ALIASES[value]


def unit_scale(unit):
    """Multiplier converting a signal in `unit` into volts."""
    return C.UNIT_SCALE[normalize_unit(unit)]


def harmonise_units(trials, unit):
    """Convert a subject's signal to volts.

    Apply to the time series, before any spectral estimate.
    """
    return np.asarray(trials, dtype=np.float64) * unit_scale(unit)


def log_power_offset(unit, half=False):
    """Constant that harmonising removes from a log-band-power feature.

    Inputs
        half  match the 0.5 factor used by differential_entropy()

    Returns
        float; positive means the unharmonised feature is that much larger.
    """
    # TODO: (@takashi) Justify the factor of 2. Amplitude scales by
    # unit_scale(unit); why does the log-power shift use twice its log?
    shift = -2.0 * np.log(unit_scale(unit))
    return 0.5 * shift if half else shift


# ------------------------------------------------------------------ loading

def load_processed(path, n_eeg=C.N_EEG, source_unit=None):
    """One subject's pre-processed EEG as (clips, channels, samples).

    This is the only place in the package that converts units. Pass
    `source_unit` and the array comes back in volts; leave it None and the
    array comes back as stored. Do not also call harmonise_units() on it.

    Inputs
        n_eeg        scalp channels to keep, from the front of the axis
        source_unit  unit the file is stored in, or None to leave it alone

    Returns
        (n_clip, n_eeg, n_sample) array

    Raises
        ValueError on any shape that does not match the documented layout.
    """
    with open(path, "rb") as handle:
        array = np.asarray(pickle.load(handle), dtype=np.float64)

    name = os.path.basename(path)
    if array.ndim != 3:
        raise ValueError("%s: expected 3-D (clips, channels, samples), got %s"
                         % (name, array.shape))
    if array.shape[:2] != (C.N_CLIP, C.N_CHANNEL_TOTAL):
        raise ValueError(
            "%s: expected (%d, %d, ...), got %s. If the first two axes look "
            "swapped the file is transposed -- establish which axis is which "
            "before reshaping."
            % (name, C.N_CLIP, C.N_CHANNEL_TOTAL, array.shape))
    if array.shape[2] != C.N_SAMPLE:
        raise ValueError("%s: expected %d samples, got %d"
                         % (name, C.N_SAMPLE, array.shape[2]))

    if array.shape[1] < n_eeg:
        raise ValueError("%s: requested %d EEG channels but the file has %d"
                         % (name, n_eeg, array.shape[1]))
    if source_unit is not None:
        array = harmonise_units(array, source_unit)
    return array[:, :n_eeg, :]


def load_feature_file(path, n_eeg=C.N_EEG):
    """A shipped DE or PSD array as (clips, channels, windows, bands)."""
    with open(path, "rb") as handle:
        array = np.asarray(pickle.load(handle), dtype=np.float64)
    if array.ndim != 4:
        raise ValueError("%s: expected 4-D (clips, channels, windows, bands), "
                         "got %s" % (os.path.basename(path), array.shape))
    if array.shape[0] != C.N_CLIP or array.shape[3] != C.N_BAND:
        raise ValueError("%s: expected (%d, ch, win, %d), got %s"
                         % (os.path.basename(path), C.N_CLIP, C.N_BAND,
                            array.shape))
    if not isinstance(n_eeg, (int, np.integer)) or n_eeg <= 0:
        raise ValueError("n_eeg must be a positive integer, got %r" % (n_eeg,))
    if array.shape[1] < n_eeg:
        raise ValueError("%s: requested %d EEG channels but the file has "
                         "only %d"
                         % (os.path.basename(path), n_eeg, array.shape[1]))
    if not np.isfinite(array).all():
        raise ValueError("%s: feature array contains NaN or infinity"
                         % os.path.basename(path))
    return array[:, :n_eeg]


def collapse_windows(feature, reducer="mean"):
    """Reduce the one-second windows to one value per trial.

    Inputs
        feature  (n_clip, n_channel, n_window, n_band) array
        reducer  "mean", "median" or "std"

    Returns
        (n_clip, n_channel, n_band) array
    """
    # TODO: (@takashi) Averaging the windows is a modelling choice, not a
    # fact. Name the alternative and say what it does to the example count.
    array = np.asarray(feature, dtype=np.float64)
    if reducer == "mean":
        return array.mean(axis=2)
    if reducer == "median":
        return np.median(array, axis=2)
    if reducer == "std":
        return array.std(axis=2, ddof=1)
    raise ValueError("reducer must be mean, median or std")


# ------------------------------------------------------------------ ratings

def load_rating_layout(path):
    """Read DataStructureOfBehaviouralData.xlsx into {item_name: column_index}.

    Returns
        dict, or None if the sheet cannot be interpreted. Never a guess.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    book = openpyxl.load_workbook(path, data_only=True)
    for sheet in book.worksheets:
        cells = [[("" if c is None else str(c).strip()) for c in row]
                 for row in sheet.iter_rows(values_only=True)]
        flat = [c.lower() for row in cells for c in row]
        if "valence" in flat and "arousal" in flat:
            for row in cells:
                lowered = [c.lower() for c in row]
                if "valence" in lowered and "arousal" in lowered:
                    return {name.lower(): i
                            for i, name in enumerate(row) if name}
            for col in range(max(len(r) for r in cells)):
                column = [r[col].lower() if col < len(r) else "" for r in cells]
                if "valence" in column and "arousal" in column:
                    return {name: i for i, name in
                            ((c, k) for k, c in enumerate(column)) if name}
    return None


def load_ratings(path, valence_col, arousal_col, key=None):
    """One subject's ratings.

    Returns
        (n_clip, 2) array, valence in column 0 and arousal in column 1.
    """
    from scipy.io import loadmat

    contents = loadmat(path, squeeze_me=True)
    if key is None:
        candidates = [k for k, v in contents.items()
                      if not k.startswith("__") and np.asarray(v).ndim == 2]
        if len(candidates) != 1:
            raise ValueError(
                "cannot pick the ratings variable in %s automatically "
                "(candidates: %s). Pass key= explicitly."
                % (os.path.basename(path), candidates))
        key = candidates[0]

    table = np.asarray(contents[key], dtype=np.float64)
    if table.ndim != 2:
        raise ValueError("%s[%s]: expected 2-D, got %s"
                         % (os.path.basename(path), key, table.shape))
    if table.shape[0] != C.N_CLIP:
        if table.shape[1] == C.N_CLIP:
            table = table.T
        else:
            raise ValueError("%s: expected %d clips, got %s"
                             % (os.path.basename(path), C.N_CLIP, table.shape))

    out = np.column_stack([table[:, valence_col], table[:, arousal_col]])
    finite = out[np.isfinite(out)]
    if finite.size:
        low, high = finite.min(), finite.max()
        if low < C.RATING_MIN - 1e-9 or high > C.RATING_MAX + 1e-9:
            raise ValueError(
                "%s: ratings span [%.3f, %.3f], outside the documented %g-%g "
                "scale. Wrong columns, or this is not the ratings table."
                % (os.path.basename(path), low, high, C.RATING_MIN,
                   C.RATING_MAX))
    return out


# ---------------------------------------------------------------- electrodes

def load_electrodes(path):
    """Read Electrode_Location.xlsx -> {cohort: [names in row order]}.

    The release has two cohorts whose electrode positions match but where six
    channels carry different names.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    book = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sheet in book.worksheets:
        rows = [[("" if c is None else str(c).strip()) for c in row]
                for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        header = [c.lower() for c in rows[0]]
        cohort_cols = [i for i, h in enumerate(header) if "cohort" in h]
        if cohort_cols:
            for col in cohort_cols:
                names = [r[col] for r in rows[1:] if col < len(r) and r[col]]
                if names:
                    label = rows[0][col]
                    out[label] = names
        else:
            names = [r[0] for r in rows[1:] if r and r[0]]
            if names:
                out[sheet.title] = names
    return out or None

